from typing import List, Dict, Optional, Tuple
import os
import io
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import font_manager
import jdatetime
from datetime import datetime
from PIL import Image
import matplotlib
matplotlib.use('Agg')


class ReportGenerator:
    @staticmethod
    def create_monthly_chart(tests: List[Dict]) -> Optional[bytes]:
        """Create monthly chart of glucose levels"""
        if not tests:
            return None

        try:
            # Sort tests by date
            tests_sorted = sorted(tests, key=lambda x: x['created_at'])

            # Prepare data
            dates = []
            glucose_values = []

            for test in tests_sorted:
                # Convert to Jalali date
                gregorian_date = datetime.fromisoformat(
                    test['created_at'].replace('Z', '+00:00'))
                jalali_date = jdatetime.datetime.fromgregorian(
                    datetime=gregorian_date)
                dates.append(jalali_date.strftime("%d/%m"))
                glucose_values.append(test['glucose'])

            # Create figure with better styling
            plt.style.use('seaborn-v0_8-darkgrid')
            fig, ax = plt.subplots(figsize=(12, 7))

            # Plot glucose values with gradient color
            line = ax.plot(dates, glucose_values, marker='o', linewidth=3, markersize=10,
                           color='#2E86AB', markerfacecolor='#FF6B6B', markeredgewidth=2)

            # Fill under the line
            ax.fill_between(dates, glucose_values, alpha=0.2, color='#2E86AB')

            # Add horizontal lines for ranges with better styling
            ranges = [
                (70, 'green', 'حد پایین نرمال', '--'),
                (100, 'blue', 'حد بالای نرمال ناشتا', '-.'),
                (140, 'orange', 'حد بالای نرمال', ':'),
                (200, 'red', 'حد خطر', '--')
            ]

            for value, color, label, linestyle in ranges:
                ax.axhline(y=value, color=color, linestyle=linestyle,
                           alpha=0.7, linewidth=2, label=label)

            # Customize plot with better styling
            ax.set_xlabel('📅 تاریخ (روز/ماه)', fontsize=14,
                          fontweight='bold', labelpad=15)
            ax.set_ylabel('🩸 میزان قند خون (mg/dL)', fontsize=14,
                          fontweight='bold', labelpad=15)
            ax.set_title('📊 نمودار ماهانه قند خون',
                         fontsize=16, fontweight='bold', pad=25)

            # Add grid
            ax.grid(True, alpha=0.4, linestyle='--')

            # Add legend with better positioning
            ax.legend(loc='upper right', fontsize=10,
                      framealpha=0.9, shadow=True)

            # Rotate and style date labels
            plt.xticks(rotation=45, fontsize=11)
            plt.yticks(fontsize=11)

            # Add value labels on points
            for i, (date, value) in enumerate(zip(dates, glucose_values)):
                ax.annotate(f'{value}', (date, value),
                            textcoords="offset points",
                            xytext=(0, 10),
                            ha='center',
                            fontsize=9,
                            fontweight='bold')

            # Adjust layout
            plt.tight_layout()

            # Add footer text
            fig.text(0.5, 0.01, 'ربات مدیریت قند خون | ایجاد شده با matplotlib',
                     ha='center', fontsize=10, alpha=0.7)

            # Save to bytes
            buf = io.BytesIO()
            plt.savefig(buf, format='png', dpi=200, bbox_inches='tight',
                        facecolor=fig.get_facecolor(), edgecolor='none')
            plt.close(fig)
            buf.seek(0)

            return buf.read()
        except Exception as e:
            print(f"Error creating chart: {e}")
            return None

    @staticmethod
    def create_excel_report(tests: List[Dict]) -> Optional[bytes]:
        """Create Excel report of tests"""
        if not tests:
            return None

        try:
            # Prepare data for DataFrame
            data = []
            for test in tests:
                gregorian_date = datetime.fromisoformat(
                    test['created_at'].replace('Z', '+00:00'))
                jalali_date = jdatetime.datetime.fromgregorian(
                    datetime=gregorian_date)

                data.append({
                    'شناسه': test['id'],
                    'تاریخ شمسی': test['shamsi_date'],
                    'ساعت آزمایش': test['test_time'],
                    'قند خون (mg/dL)': test['glucose'],
                    'نوع آزمایش': 'ناشتا' if test['fasting'] else 'غیرناشتا',
                    'علائم': test['symptoms'],
                    'یادداشت': test.get('notes', ''),
                    'تاریخ ثبت': gregorian_date.strftime("%Y-%m-%d %H:%M")
                })

            # Create DataFrame
            df = pd.DataFrame(data)

            # Add statistics row
            if len(df) > 0:
                stats_row = {
                    'شناسه': 'آمار',
                    'تاریخ شمسی': '',
                    'ساعت آزمایش': '',
                    'قند خون (mg/dL)': df['قند خون (mg/dL)'].mean(),
                    'نوع آزمایش': '',
                    'علائم': f"تعداد: {len(df)} | حداقل: {df['قند خون (mg/dL)'].min()} | حداکثر: {df['قند خون (mg/dL)'].max()}",
                    'یادداشت': '',
                    'تاریخ ثبت': ''
                }
                df = pd.concat([df, pd.DataFrame([stats_row])],
                               ignore_index=True)

            # Create Excel file in memory with styling
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(
                    writer, sheet_name='آزمایش‌های قند خون', index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['آزمایش‌های قند خون']

                # Style the header
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                # Define styles
                header_font = Font(name='Arial', bold=True,
                                   size=12, color='FFFFFF')
                header_fill = PatternFill(
                    start_color='2E86AB', end_color='2E86AB', fill_type='solid')
                cell_alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

                # Apply styles to header
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = cell_alignment

                # Auto-adjust columns width
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 4, 40)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Style the statistics row
                if len(df) > 0:
                    stats_row_num = len(df) + 1
                    stats_fill = PatternFill(
                        start_color='FFEAA7', end_color='FFEAA7', fill_type='solid')
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=stats_row_num, column=col)
                        cell.fill = stats_fill
                        cell.font = Font(bold=True)

            output.seek(0)
            return output.read()
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            return None

    @staticmethod
    def create_pdf_report(tests: List[Dict]) -> Optional[bytes]:
        """Create PDF report of tests (returns image for now, can be extended to actual PDF)"""
        if not tests:
            return None

        try:
            # Create a styled text report as image
            from PIL import Image, ImageDraw, ImageFont

            # Create image
            img = Image.new('RGB', (800, 1200), color=(255, 255, 255))
            draw = ImageDraw.Draw(img)

            # Try to load font
            try:
                font_large = ImageFont.truetype("arial.ttf", 24)
                font_medium = ImageFont.truetype("arial.ttf", 18)
                font_small = ImageFont.truetype("arial.ttf", 14)
            except:
                font_large = ImageFont.load_default()
                font_medium = ImageFont.load_default()
                font_small = ImageFont.load_default()

            # Draw header
            draw.text((400, 50), "گزارش آزمایش‌های قند خون",
                      font=font_large, fill=(0, 0, 0), anchor="mm")

            # Draw statistics
            glucose_values = [t['glucose'] for t in tests]
            stats_text = f"""
            آمار کلی:
            • تعداد آزمایش‌ها: {len(tests)}
            • میانگین قند خون: {sum(glucose_values)/len(glucose_values):.1f} mg/dL
            • حداقل: {min(glucose_values)} mg/dL
            • حداکثر: {max(glucose_values)} mg/dL
            """

            draw.text((400, 150), stats_text, font=font_medium,
                      fill=(0, 0, 0), anchor="mm", align="center")

            # Draw table header
            y = 250
            headers = ["تاریخ", "ساعت", "قند خون", "نوع", "علائم"]
            for i, header in enumerate(headers):
                draw.text((50 + i * 150, y), header, font=font_medium,
                          fill=(46, 134, 171))

            # Draw separator line
            draw.line([(50, y + 30), (750, y + 30)], fill=(0, 0, 0), width=2)

            # Draw test data
            y += 50
            for test in tests[:15]:  # Limit to 15 rows
                draw.text((50, y), test['shamsi_date'],
                          font=font_small, fill=(0, 0, 0))
                draw.text((200, y), test['test_time'],
                          font=font_small, fill=(0, 0, 0))
                draw.text((350, y), str(test['glucose']),
                          font=font_small, fill=(0, 0, 0))
                draw.text((500, y), "ناشتا" if test['fasting'] else "غیرناشتا",
                          font=font_small, fill=(0, 0, 0))
                draw.text((650, y), test['symptoms'][:15],
                          font=font_small, fill=(0, 0, 0))
                y += 40

            # Save to bytes
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            buf.seek(0)

            return buf.read()
        except Exception as e:
            print(f"Error creating PDF/image report: {e}")
            return None

    @staticmethod
    def create_text_report(tests: List[Dict], report_type: str = "هفتگی") -> str:
        """Create formatted text report of tests"""
        if not tests:
            return f"❌ هیچ آزمایشی برای گزارش {report_type} یافت نشد."

        try:
            report = "📊 " + "="*40 + "\n"
            report += f"گزارش {report_type} آزمایش‌های قند خون\n"
            report += "="*40 + "\n\n"

            # Calculate statistics
            glucose_values = [t['glucose'] for t in tests]
            avg_glucose = sum(glucose_values) / len(glucose_values)
            fasting_count = len([t for t in tests if t['fasting']])
            non_fasting_count = len(tests) - fasting_count

            # Add statistics
            report += "📈 آمار کلی:\n"
            report += "─"*30 + "\n"
            report += f"• تعداد کل آزمایش‌ها: {len(tests)} عدد\n"
            report += f"• میانگین قند خون: {avg_glucose:.1f} mg/dL\n"
            report += f"• حداقل مقدار: {min(glucose_values)} mg/dL\n"
            report += f"• حداکثر مقدار: {max(glucose_values)} mg/dL\n"
            report += f"• آزمایش‌های ناشتا: {fasting_count} عدد\n"
            report += f"• آزمایش‌های غیرناشتا: {non_fasting_count} عدد\n\n"

            # Add individual tests
            report += "📋 لیست آزمایش‌ها:\n"
            report += "─"*30 + "\n"

            for i, test in enumerate(tests[:10], 1):  # Limit to 10 tests
                status_emoji = "🟢" if test['glucose'] <= 140 else "🟡" if test['glucose'] <= 200 else "🔴"
                fasting_emoji = "🟦" if test['fasting'] else "🟧"

                report += f"{i}. {status_emoji} {test['shamsi_date']} - ساعت {test['test_time']}\n"
                report += f"   مقدار: {test['glucose']} mg/dL | نوع: {fasting_emoji} "
                report += "ناشتا" if test['fasting'] else "غیرناشتا"
                report += f"\n   علائم: {test['symptoms']}\n"

                if test.get('notes'):
                    report += f"   📝 یادداشت: {test['notes']}\n"

                report += "\n"

            if len(tests) > 10:
                report += f"... و {len(tests) - 10} آزمایش دیگر\n\n"

            report += "📅 تاریخ گزارش: " + jdatetime.datetime.now().strftime("%Y/%m/%d %H:%M")
            report += "\n" + "="*40 + "\n"

            return report
        except Exception as e:
            print(f"Error creating text report: {e}")
            return f"❌ خطا در ایجاد گزارش: {str(e)}"


# Create global report generator instance
report_generator = ReportGenerator()
